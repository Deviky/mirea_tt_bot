import openpyxl

def parse(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)

    # 1) groups
    groups_sheet = wb['Groups']
    groups = []
    for row in groups_sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        group_id, year = row
        if group_id is None:
            continue
        groups.append({
            "group_id": group_id,
            "year": year
        })

    # 2) teachers
    teachers_sheet = wb['Teachers']
    teachers = []
    for row in teachers_sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        full_name, department = row
        if full_name is None:
            continue
        teachers.append({
            "full_name": full_name,
            "department": department
        })

    # 3) time_table_info
    time_table_info = []

    # Список дней недели для поиска, по твоему примеру
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

    # Парсим каждый лист, кроме Groups и Teachers
    group_sheets = [name for name in wb.sheetnames if name not in ('Groups', 'Teachers')]

    for sheet_name in group_sheets:
        sheet = wb[sheet_name]

        current_day = None

        # Пройдемся по строкам листа
        # В твоем примере день недели в первой колонке (A),
        # пары идут по строкам, где в колонке A написано "1 пара", "2 пара" и т.д.
        for row in sheet.iter_rows(min_row=1, values_only=True):
            first_cell = row[0]
            if first_cell in days_of_week:
                current_day = first_cell
                continue
            if current_day is None:
                # еще не дошли до расписания
                continue

            # Теперь ожидаем, что в первой колонке будет что-то типа "1 пара", "2 пара" и т.д.
            if first_cell is None:
                continue
            if "пара" not in str(first_cell):
                continue

            # номер пары
            couple_num_str = str(first_cell).strip()
            # извлечь число
            try:
                couple_num = int(couple_num_str.split()[0])
            except:
                continue

            # По формату:
            # Колонки:
            # A: "1 пара" (couple_num)
            # B: Предмет 1 неделя
            # C: Аудитория 1 неделя
            # D: Преподаватель 1 неделя
            # E: Предмет 2 неделя
            # F: Аудитория 2 неделя
            # G: Преподаватель 2 неделя

            # Проверим что колонок достаточно
            if len(row) < 7:
                continue

            # Неделя 1
            course_1 = row[1]
            auditorium_1 = row[2]
            teacher_1 = row[3]

            if course_1 and teacher_1:
                time_table_info.append({
                    "group_id": sheet_name,
                    "week_num": 1,
                    "day_of_week": current_day,
                    "couple_num": couple_num,
                    "course_name": course_1,
                    "auditorium": auditorium_1,
                    "teacher_full_name": teacher_1
                })

            # Неделя 2
            course_2 = row[4]
            auditorium_2 = row[5]
            teacher_2 = row[6]

            if course_2 and teacher_2:
                time_table_info.append({
                    "group_id": sheet_name,
                    "week_num": 2,
                    "day_of_week": current_day,
                    "couple_num": couple_num,
                    "course_name": course_2,
                    "auditorium": auditorium_2,
                    "teacher_full_name": teacher_2
                })

    return groups, teachers, time_table_info